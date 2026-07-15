#!/usr/bin/env python3
"""
processar_dec_abril_16mm.py
----------------------------
Calcula o primeiro dia (a partir de 1º de abril) em que a precipitação
acumulada atinge >= 16 mm para cada município-ano em uma planilha Excel.

Usa dados diários de precipitação do BR-DWGD v3.2.4, com ponderação
por área de interseção pixel (0.1° × 0.1°) × município.

Uso:
    python processar_dec_abril_16mm.py \
        --planilha PAM_CAFE_MATAS_clima.xlsx \
        --sheet "DELTA PRODUTIVIDADE (brutos)" \
        --gpkg MUN_MATAS.gpkg \
        --coluna_mun Municipio \
        --coluna_ano Ano \
        --coluna_cod "Código IBGE" \
        --coluna_alvo Dec_abril_16mm \
        --dir_dados /tmp/pr_daily \
        --limiar 16.0

Citação obrigatória dos dados:
    Xavier, A. C. et al. (2022) New improved Brazilian daily weather gridded
    data (1961–2020). Int. J. Climatol., 42(16), 8390–8404.
    https://doi.org/10.1002/joc.7731
"""

import argparse, os, sys, time, warnings
import numpy as np
import pandas as pd
import xarray as xr
import geopandas as gpd
import openpyxl
from shapely.geometry import box
from datetime import datetime

warnings.filterwarnings("ignore")

# ===========================================================================
# Constantes
# ===========================================================================
CRS_GEO = "EPSG:4674"       # SIRGAS 2000 (lat/lon)
CRS_PLANAR = "EPSG:5880"    # Policônica do Brasil (para áreas)

# Arquivos NetCDF diários do BR-DWGD v3.2.4
ARQUIVOS_PADRAO_NC = [
    "pr_19610101_19801231_BR-DWGD_UFES_UTEXAS_v_3.2.4.nc",
    "pr_19810101_20001231_BR-DWGD_UFES_UTEXAS_v_3.2.4.nc",
    "pr_20010101_20251231_BR-DWGD_UFES_UTEXAS_v_3.2.4.nc",
]


# ===========================================================================
# Funções auxiliares
# ===========================================================================

def compute_pixel_weights(municipio_polygon, lats, lons):
    """
    Calcula a matriz de pesos para cada pixel da grade (lat, lon)
    com base na área de interseção com o polígono do município.

    Retorna:
        weight_map : ndarray shape (len(lats)-1, len(lons)-1)
                     peso de cada pixel (soma = 1.0)
        pixel_indices : lista de tuplas (i, j) com pixels que intersectam
    """
    pixel_polys = []
    pixel_indices = []
    n_lat = len(lats)
    n_lon = len(lons)

    for i in range(n_lat - 1):
        for j in range(n_lon - 1):
            cell = box(
                float(lons[j]), float(lats[i + 1]),   # left, bottom
                float(lons[j + 1]), float(lats[i])     # right, top
            )
            if cell.intersects(municipio_polygon):
                pixel_polys.append(cell)
                pixel_indices.append((i, j))

    if not pixel_polys:
        return None, []

    # Projetar para cálculo de área
    gdf_pixels = gpd.GeoDataFrame(
        {"value": 1.0, "geometry": pixel_polys}, crs=CRS_GEO
    ).to_crs(CRS_PLANAR)

    gdf_mun = gpd.GeoDataFrame(
        {"geometry": [municipio_polygon]}, crs=CRS_GEO
    ).to_crs(CRS_PLANAR)

    mun_geom_planar = gdf_mun.geometry.iloc[0]
    areas_inter = gdf_pixels.intersection(mun_geom_planar).area.values

    total_area = areas_inter.sum()
    if total_area == 0:
        return None, []

    # Normalizar pesos
    weights = areas_inter / total_area

    # Construir mapa de pesos esparso
    weight_map = {}
    for (i, j), w in zip(pixel_indices, weights):
        weight_map[(i, j)] = w

    return weight_map, pixel_indices


def primeira_data_acumulado_16mm(pr_municipio, time_index, ano, limiar=16.0):
    """
    Encontra a primeira data a partir de 1º de abril do ano em que
    a precipitação acumulada atinge >= limiar mm.

    Args:
        pr_municipio: array 1D de precipitação diária
        time_index: pd.DatetimeIndex correspondente
        ano: ano de interesse
        limiar: limiar de precipitação acumulada (mm)

    Returns:
        dict com 'data', 'dias_desde_abril', 'pr_acum', ou None
    """
    data_ini = pd.Timestamp(f"{ano}-04-01")
    data_fim = pd.Timestamp(f"{ano}-12-31")

    mask = (time_index >= data_ini) & (time_index <= data_fim)
    idx_periodo = np.where(mask)[0]

    if len(idx_periodo) == 0:
        return None

    pr_periodo = pr_municipio[idx_periodo]
    pr_acum = np.cumsum(pr_periodo)

    idx_atinge = np.where(pr_acum >= limiar)[0]

    if len(idx_atinge) == 0:
        return {"data": None, "dias_desde_abril": None, "pr_acum": None,
                "status": "NAO_ATINGIU"}

    dia = idx_atinge[0]
    return {
        "data": time_index[idx_periodo[dia]],
        "dias_desde_abril": dia,
        "pr_acum": float(pr_acum[dia]),
        "status": "ATINGIU",
    }


# ===========================================================================
# Pipeline principal
# ===========================================================================

def processar(
    planilha_path,
    sheet_name,
    gpkg_path,
    col_municipio,
    col_ano,
    col_codigo,
    col_alvo,
    dir_dados,
    limiar=16.0,
    saida_path=None,
):
    """
    Pipeline completo:
    1. Lê planilha e geometrias
    2. Carrega dados BR-DWGD (subset regional)
    3. Calcula pesos de área por município
    4. Para cada município-ano, encontra o primeiro dia com precip. acum. >= limiar
    5. Escreve resultados na planilha
    """
    t0 = time.time()

    # ── 1. Planilha ──
    print("[1] Lendo planilha...")
    wb = openpyxl.load_workbook(planilha_path, data_only=True)
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]

    # Mapear colunas
    try:
        idx_mun = headers.index(col_municipio)
        idx_ano = headers.index(col_ano)
        idx_cod = headers.index(col_codigo)
        idx_alvo = headers.index(col_alvo)
    except ValueError as e:
        raise ValueError(f"Coluna não encontrada no header: {e}")

    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None and row[2] is None:
            continue
        if all(v is None for v in row[:max(idx_mun, idx_ano, idx_cod) + 1]):
            continue
        rows_data.append(row)

    print(f"   {len(rows_data)} linhas de dados")

    # Dicionário: (município, ano) → índice na lista
    mun_anos = {}
    codigos = {}
    for i, row in enumerate(rows_data):
        mun = str(row[idx_mun]).strip() if row[idx_mun] else None
        ano = int(row[idx_ano]) if row[idx_ano] else None
        cod = row[idx_cod]
        if mun and ano:
            mun_anos[(mun, ano)] = i
            codigos[mun] = cod

    print(f"   {len(mun_anos)} municípios-ano únicos")

    # ── 2. Geometrias ──
    print("[2] Carregando geometrias...")
    gdf = gpd.read_file(gpkg_path)
    # Garantir CD_MUN como string
    if "CD_MUN" in gdf.columns:
        gdf["CD_MUN"] = gdf["CD_MUN"].astype(str)
        gdf = gdf.set_index("CD_MUN")

    mun_to_cod = {}
    for mun, cod in codigos.items():
        if cod is not None:
            cod_str = str(int(cod))
            mun_to_cod[mun] = cod_str

    print(f"   {len(mun_to_cod)} municípios mapeados para geometrias")

    # Bounding box para subset
    bounds = gdf.total_bounds
    lon_min_b, lat_min_b, lon_max_b, lat_max_b = bounds
    # Margem de 2 pixels
    MARGEM = 0.2
    LON_MIN = lon_min_b - MARGEM
    LON_MAX = lon_max_b + MARGEM
    LAT_MIN = lat_min_b - MARGEM
    LAT_MAX = lat_max_b + MARGEM

    # ── 3. Dados BR-DWGD ──
    print("[3] Carregando dados BR-DWGD (subset)...")
    ds_list = []
    for nc_name in ARQUIVOS_PADRAO_NC:
        nc_path = os.path.join(dir_dados, nc_name)
        if not os.path.exists(nc_path):
            print(f"   ⚠ {nc_name} não encontrado — pulando")
            continue
        print(f"   Abrindo {nc_name}...")
        ds = xr.open_dataset(nc_path, chunks={"time": 365})

        # Renomear coordenadas se necessário
        ds = ds.rename(
            {k: v for k, v in
             [("lat", "latitude"), ("lon", "longitude"),
              ("Lat", "latitude"), ("Lon", "longitude")]
             if k in ds.dims}
        )

        # Latitude aumenta de sul para norte (v3.2.4)
        # Lon aumenta de oeste para leste
        ds_sub = ds.sel(
            latitude=slice(LAT_MIN, LAT_MAX),
            longitude=slice(LON_MIN, LON_MAX),
        )
        ds_sub = ds_sub.load()
        ds_list.append(ds_sub)
        print(f"      → {len(ds_sub.time)} dias, "
              f"{len(ds_sub.latitude)} lat × {len(ds_sub.longitude)} lon")

    if not ds_list:
        raise RuntimeError("Nenhum arquivo NetCDF foi carregado!")

    ds_all = xr.concat(ds_list, dim="time") if len(ds_list) > 1 else ds_list[0]

    # Remover duplicatas temporais
    _, uniq_idx = np.unique(ds_all.time.values, return_index=True)
    if len(uniq_idx) < len(ds_all.time):
        print(f"   Removendo {len(ds_all.time) - len(uniq_idx)} timesteps duplicados")
        ds_all = ds_all.isel(time=np.sort(uniq_idx))

    lats = ds_all.latitude.values
    lons = ds_all.longitude.values
    times = ds_all.time.values
    pr_data = ds_all.pr.values  # (time, lat, lon)
    time_idx = pd.DatetimeIndex(times)

    # Verificar resolução
    lat_res = abs(np.mean(np.diff(lats)))
    lon_res = abs(np.mean(np.diff(lons)))
    assert abs(lat_res - 0.1) < 1e-4, f"Resolução lat {lat_res} ≠ 0.1°"
    assert abs(lon_res - 0.1) < 1e-4, f"Resolução lon {lon_res} ≠ 0.1°"

    print(f"   Grade: {len(lats)} lat × {len(lons)} lon | Res: {lat_res:.2f}°")
    print(f"   Período: {time_idx[0].strftime('%Y-%m-%d')} a {time_idx[-1].strftime('%Y-%m-%d')}")
    print(f"   Memória: {pr_data.nbytes / 1e6:.1f} MB | Tempo: {time.time()-t0:.1f}s")

    # ── 4. Pesos por município ──
    print("[4] Calculando pesos de área...")
    municipios_ok = {}
    sem_geo = []
    sem_inter = []

    for mun_nome, cod_str in mun_to_cod.items():
        if cod_str not in gdf.index:
            sem_geo.append(mun_nome)
            continue

        poly = gdf.loc[cod_str, "geometry"]
        weight_map, px_indices = compute_pixel_weights(poly, lats, lons)

        if weight_map is None:
            sem_inter.append(mun_nome)
            continue

        municipios_ok[mun_nome] = {
            "codigo": cod_str,
            "pesos": weight_map,
            "pixels": px_indices,
        }

    print(f"   {len(municipios_ok)} municípios com pesos calculados")
    if sem_geo:
        print(f"   ⚠ Sem geometria: {sem_geo}")
    if sem_inter:
        print(f"   ⚠ Sem interseção: {sem_inter}")

    # ── 5. Calcular para cada município-ano ──
    print(f"[5] Calculando Dec_abril_{limiar:.0f}mm para {len(mun_anos)} municípios-ano...")
    resultados = {}

    for mun_nome, info in municipios_ok.items():
        weight_map = info["pesos"]

        # Série temporal do município
        pr_mun = np.zeros(len(times), dtype=np.float64)
        for (i, j), w in weight_map.items():
            pr_mun += pr_data[:, i, j] * w

        # Processar anos deste município
        for (m, a) in [(m, a) for (m, a) in mun_anos.keys() if m == mun_nome]:
            res = primeira_data_acumulado_16mm(pr_mun, time_idx, a, limiar)
            resultados[(mun_nome, a)] = res

    print(f"   Concluído em {time.time() - t0:.1f}s")

    # ── 6. Escrever na planilha ──
    print("[6] Salvando resultados na planilha...")
    wb_out = openpyxl.load_workbook(planilha_path)
    ws_out = wb_out[sheet_name]

    # Encontrar letra da coluna
    col_alvo_letra = None
    for cell in ws_out[1]:
        if cell.value == col_alvo:
            col_alvo_letra = cell.column_letter
            break
    if col_alvo_letra is None:
        raise RuntimeError(f"Coluna '{col_alvo}' não encontrada no header")

    cont_atingiu = 0
    cont_nao = 0
    cont_sem = 0

    for i, row in enumerate(rows_data):
        mun = str(row[idx_mun]).strip() if row[idx_mun] else None
        ano = int(row[idx_ano]) if row[idx_ano] else None
        if mun is None or ano is None:
            continue

        linha_excel = i + 2  # 1-based, pulando header
        chave = (mun, ano)

        if chave in resultados:
            res = resultados[chave]
            if res and res["data"] is not None:
                valor = res["data"].strftime("%d/%m/%Y")
                cont_atingiu += 1
            elif res and res["status"] == "NAO_ATINGIU":
                valor = "NÃO ATINGIU"
                cont_nao += 1
            else:
                valor = "SEM DADOS"
                cont_sem += 1
        else:
            if mun in sem_geo:
                valor = "SEM GEOMETRIA"
            elif mun in sem_inter:
                valor = "SEM INTERSECAO"
            else:
                valor = "ERRO"

        ws_out[f"{col_alvo_letra}{linha_excel}"] = valor

    # Salvar
    if saida_path is None:
        saida_path = planilha_path
    wb_out.save(saida_path)

    # ── 7. Resumo ──
    print(f"\n{'=' * 60}")
    print("RESUMO")
    print(f"{'=' * 60}")
    print(f"  Total municípios-ano: {len(mun_anos)}")
    print(f"  ✅ Data encontrada:     {cont_atingiu}")
    print(f"  ❌ Não atingiu {limiar}mm: {cont_nao}")
    print(f"  ⚠  Sem dados:           {cont_sem}")
    print(f"  🌍 Municípios OK:       {len(municipios_ok)}/{len(mun_to_cod)}")
    print(f"\n  Arquivo salvo: {saida_path}")
    print(f"  Tempo total: {time.time() - t0:.1f}s")
    print(f"  Coluna: {col_alvo_letra} ({col_alvo})")

    return resultados


# ===========================================================================
# CLI
# ===========================================================================

def parse_args():
    p = argparse.ArgumentParser(
        description="Calcula o 1º dia (a partir de abril) com precipitação acumulada >= limiar.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplo:
  python processar_dec_abril_16mm.py \\
      --planilha PAM_CAFE_MATAS_clima.xlsx \\
      --sheet "DELTA PRODUTIVIDADE (brutos)" \\
      --gpkg municipios.gpkg \\
      --dir_dados /tmp/pr_daily \\
      --limiar 16.0
        """,
    )
    p.add_argument("--planilha", required=True, help="Caminho da planilha Excel (.xlsx)")
    p.add_argument("--sheet", default="DELTA PRODUTIVIDADE (brutos)",
                   help="Nome da aba na planilha")
    p.add_argument("--gpkg", required=True, help="GeoPackage com geometrias dos municípios")
    p.add_argument("--dir_dados", default="/tmp/pr_daily",
                   help="Diretório com arquivos NetCDF diários do BR-DWGD")
    p.add_argument("--coluna_mun", default="Município", help="Nome da coluna do município")
    p.add_argument("--coluna_ano", default="Ano", help="Nome da coluna do ano")
    p.add_argument("--coluna_cod", default="Código IBGE", help="Nome da coluna do código IBGE")
    p.add_argument("--coluna_alvo", default="Dec_abril_16mm",
                   help="Nome da coluna alvo para escrita")
    p.add_argument("--limiar", type=float, default=16.0,
                   help="Limiar de precipitação acumulada em mm (padrão: 16.0)")
    p.add_argument("--saida", default=None,
                   help="Caminho de saída (padrão: sobrescreve a planilha de entrada)")
    return p.parse_args()


def main():
    args = parse_args()
    processar(
        planilha_path=args.planilha,
        sheet_name=args.sheet,
        gpkg_path=args.gpkg,
        col_municipio=args.coluna_mun,
        col_ano=args.coluna_ano,
        col_codigo=args.coluna_cod,
        col_alvo=args.coluna_alvo,
        dir_dados=args.dir_dados,
        limiar=args.limiar,
        saida_path=args.saida,
    )


if __name__ == "__main__":
    main()
