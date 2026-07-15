#!/usr/bin/env python3
"""
calcular_antese_v2.py
---------------------
Calcula a data em que o somatório de graus-dia (GD) atinge um limiar
pré-definido a partir de uma data inicial (ex: Dec_abril_16mm) para cada
município-ano em uma planilha Excel.

Grau-dia diário:
    GD = (Tmax + Tmin) / 2 - Tbase

    onde Tbase = 8.5°C (padrão para café arábica)
    Se GD < 0, GD = 0 (não acumula valores negativos)

Usa dados diários de Tmax e Tmin do BR-DWGD v3.2.4, com ponderação
por área de interseção pixel (0.1° × 0.1°) × município.

Uso:
    python calcular_antese_v2.py \
        --planilha PAM_CAFE_MATAS_clima.xlsx \
        --sheet "DELTA PRODUTIVIDADE (brutos)" \
        --gpkg MUN_MATAS.gpkg \
        --dir_dados /tmp/br-dwgd-daily \
        --coluna_mun "Município" \
        --coluna_ano "Ano" \
        --coluna_cod "Código IBGE" \
        --coluna_dec "Dec_abril_16mm" \
        --coluna_antese_v2 "Antese_v2" \
        --soma_alvo 1980.0 \
        --base_temp 8.5

Citação obrigatória dos dados:
    Xavier, A. C. et al. (2022) New improved Brazilian daily weather gridded
    data (1961–2020). Int. J. Climatol., 42(16), 8390–8404.
    https://doi.org/10.1002/joc.7731
"""

import argparse, os, sys, time, warnings
import numpy as np
import pandas as pd
import xarray as xr
import geopandas as gpd
import openpyxl
from shapely.geometry import box

warnings.filterwarnings("ignore")

# ===========================================================================
# Constantes
# ===========================================================================
CRS_GEO = "EPSG:4674"       # SIRGAS 2000 (lat/lon)
CRS_PLANAR = "EPSG:5880"    # Policônica do Brasil (para áreas)
SOMA_ALVO_PADRAO = 1980.0   # °C-dia alvo (padrão para café)
BASE_TEMP_PADRAO = 8.5      # °C - temperatura base

# Arquivos NetCDF diários do BR-DWGD v3.2.4
ARQUIVOS_TMAX = [
    "Tmax_19610101_19801231_BR-DWGD_UFES_UTEXAS_v_3.2.4.nc",
    "Tmax_19810101_20001231_BR-DWGD_UFES_UTEXAS_v_3.2.4.nc",
    "Tmax_20010101_20251231_BR-DWGD_UFES_UTEXAS_v_3.2.4.nc",
]
ARQUIVOS_TMIN = [
    "Tmin_19610101_19801231_BR-DWGD_UFES_UTEXAS_v_3.2.4.nc",
    "Tmin_19810101_20001231_BR-DWGD_UFES_UTEXAS_v_3.2.4.nc",
    "Tmin_20010101_20251231_BR-DWGD_UFES_UTEXAS_v_3.2.4.nc",
]


# ===========================================================================
# Funções principais
# ===========================================================================

def compute_pixel_weights(municipio_polygon, lats, lons):
    """
    Calcula a matriz de pesos para cada pixel da grade (lat, lon)
    com base na área de interseção com o polígono do município.

    Retorna:
        weight_map : dict {(i,j): peso} — soma dos pesos = 1.0
        pixel_indices : list de tuplas (i, j)
    """
    pixel_polys = []
    pixel_indices = []
    n_lat = len(lats)
    n_lon = len(lons)

    for i in range(n_lat - 1):
        for j in range(n_lon - 1):
            cell = box(
                float(lons[j]), float(lats[i + 1]),   # left, bottom
                float(lons[j + 1]), float(lats[i])     # right, top
            )
            if cell.intersects(municipio_polygon):
                pixel_polys.append(cell)
                pixel_indices.append((i, j))

    if not pixel_polys:
        return None, []

    # Projetar para cálculo de área planar
    gdf_pixels = gpd.GeoDataFrame(
        {"value": 1.0, "geometry": pixel_polys}, crs=CRS_GEO
    ).to_crs(CRS_PLANAR)

    gdf_mun = gpd.GeoDataFrame(
        {"geometry": [municipio_polygon]}, crs=CRS_GEO
    ).to_crs(CRS_PLANAR)

    mun_geom_planar = gdf_mun.geometry.iloc[0]
    areas_inter = gdf_pixels.intersection(mun_geom_planar).area.values

    total_area = areas_inter.sum()
    if total_area == 0:
        return None, []

    # Normalizar pesos
    weights = areas_inter / total_area

    weight_map = {}
    for (i, j), w in zip(pixel_indices, weights):
        weight_map[(i, j)] = w

    return weight_map, pixel_indices


def carregar_dados_temperatura(dir_dados, lats, lons):
    """
    Carrega e concatena os dados de Tmax e Tmin para a região de interesse.

    Args:
        dir_dados: diretório com os arquivos NetCDF
        lats: array com [lat_min, lat_max] para subset
        lons: array com [lon_min, lon_max] para subset

    Retorna:
        tmax_all: ndarray (time, lat, lon)
        tmin_all: ndarray (time, lat, lon)
        time_idx: pd.DatetimeIndex
        lats_out: array de latitudes do subset
        lons_out: array de longitudes do subset
    """
    LAT_MIN, LAT_MAX = float(lats.min()), float(lats.max())
    LON_MIN, LON_MAX = float(lons.min()), float(lons.max())

    print(f"  Subset espacial: lon=[{LON_MIN:.4f}, {LON_MAX:.4f}], "
          f"lat=[{LAT_MIN:.4f}, {LAT_MAX:.4f}]")

    # Carregar Tmax
    print("  Carregando Tmax...")
    ds_tmax_list = []
    for nc_name in ARQUIVOS_TMAX:
        nc_path = os.path.join(dir_dados, nc_name)
        if not os.path.exists(nc_path):
            print(f"    ⚠ {nc_name} não encontrado — pulando")
            continue
        ds = xr.open_dataset(nc_path, chunks={"time": 365})
        for old, new in [("lat", "latitude"), ("lon", "longitude")]:
            if old in ds.dims:
                ds = ds.rename({old: new})
        ds_sub = ds.sel(
            latitude=slice(LAT_MIN, LAT_MAX),
            longitude=slice(LON_MIN, LON_MAX),
        )
        ds_sub = ds_sub.load()
        ds_tmax_list.append(ds_sub)
        ds.close()

    # Carregar Tmin
    print("  Carregando Tmin...")
    ds_tmin_list = []
    for nc_name in ARQUIVOS_TMIN:
        nc_path = os.path.join(dir_dados, nc_name)
        if not os.path.exists(nc_path):
            continue
        ds = xr.open_dataset(nc_path, chunks={"time": 365})
        for old, new in [("lat", "latitude"), ("lon", "longitude")]:
            if old in ds.dims:
                ds = ds.rename({old: new})
        ds_sub = ds.sel(
            latitude=slice(LAT_MIN, LAT_MAX),
            longitude=slice(LON_MIN, LON_MAX),
        )
        ds_sub = ds_sub.load()
        ds_tmin_list.append(ds_sub)
        ds.close()

    # Concatenar períodos
    ds_tmax_all = xr.concat(ds_tmax_list, dim="time") if len(ds_tmax_list) > 1 else ds_tmax_list[0]
    ds_tmin_all = xr.concat(ds_tmin_list, dim="time") if len(ds_tmin_list) > 1 else ds_tmin_list[0]

    # Remover duplicatas temporais (overlap entre arquivos)
    _, uniq_idx = np.unique(ds_tmax_all.time.values, return_index=True)
    if len(uniq_idx) < len(ds_tmax_all.time):
        print(f"  Removendo {len(ds_tmax_all.time) - len(uniq_idx)} timesteps duplicados")
        ds_tmax_all = ds_tmax_all.isel(time=np.sort(uniq_idx))
        ds_tmin_all = ds_tmin_all.isel(time=np.sort(uniq_idx))

    tmax_all = ds_tmax_all["Tmax"].values
    tmin_all = ds_tmin_all["Tmin"].values
    time_idx = pd.DatetimeIndex(ds_tmax_all.time.values)
    lats_out = ds_tmax_all.latitude.values
    lons_out = ds_tmax_all.longitude.values

    print(f"  Dados carregados: {len(time_idx)} dias, "
          f"{len(lats_out)} lat × {len(lons_out)} lon")
    print(f"  Período: {time_idx[0].strftime('%Y-%m-%d')} a "
          f"{time_idx[-1].strftime('%Y-%m-%d')}")
    print(f"  Memória Tmax: {tmax_all.nbytes / 1e6:.1f} MB")
    print(f"  Memória Tmin: {tmin_all.nbytes / 1e6:.1f} MB")

    return tmax_all, tmin_all, time_idx, lats_out, lons_out


def calcular_grau_dia_diario(tmax_mun, tmin_mun, base_temp=BASE_TEMP_PADRAO):
    """
    Calcula graus-dia diários a partir de Tmax e Tmin.

    GD = (Tmax + Tmin) / 2 - base_temp
    Se GD < 0, GD = 0 (não acumula valores negativos).

    Args:
        tmax_mun: array 1D de Tmax diária
        tmin_mun: array 1D de Tmin diária
        base_temp: temperatura base (°C)

    Returns:
        gd: array 1D de graus-dia diários (valores >= 0)
    """
    tmed = (tmax_mun + tmin_mun) / 2.0
    gd = tmed - base_temp
    gd = np.maximum(gd, 0.0)
    return gd


def encontrar_data_atinge_soma(gd_array, time_idx, data_inicio,
                                soma_alvo=SOMA_ALVO_PADRAO):
    """
    Encontra a primeira data a partir de data_inicio em que a soma
    acumulada de graus-dia atinge soma_alvo.

    Args:
        gd_array: array 1D de graus-dia diários
        time_idx: pd.DatetimeIndex completo
        data_inicio: pd.Timestamp da data de partida
        soma_alvo: valor alvo de soma (°C-dia)

    Returns:
        (data_atinge, soma_atingiu, dias_corridos):
            data_atinge: pd.Timestamp ou None
            soma_atingiu: float (soma acumulada na data)
            dias_corridos: int (dias desde data_inicio)
    """
    mask_inicio = time_idx >= data_inicio
    idx_inicio = np.where(mask_inicio)[0]

    if len(idx_inicio) == 0:
        return None, 0, 0

    i0 = idx_inicio[0]
    gd_periodo = gd_array[i0:]
    soma_acum = np.cumsum(gd_periodo)

    idx_alvo = np.where(soma_acum >= soma_alvo)[0]

    if len(idx_alvo) == 0:
        return None, float(soma_acum[-1]), len(gd_periodo)

    i_alvo = idx_alvo[0]
    data_atinge = time_idx[i0 + i_alvo]
    soma_atingiu = float(soma_acum[i_alvo])
    dias_corridos = i_alvo

    return data_atinge, soma_atingiu, dias_corridos


# ===========================================================================
# Pipeline principal
# ===========================================================================

def processar(
    planilha_path,
    sheet_name,
    gpkg_path,
    dir_dados,
    col_municipio="Município",
    col_ano="Ano",
    col_codigo="Código IBGE",
    col_dec="Dec_abril_16mm",
    col_antese_v2="Antese_v2",
    soma_alvo=SOMA_ALVO_PADRAO,
    base_temp=BASE_TEMP_PADRAO,
):
    """
    Pipeline completo para cálculo da Antese_v2 (soma térmica).

    1. Lê planilha e geometrias municipais
    2. Carrega dados BR-DWGD (Tmax, Tmin, subset regional)
    3. Calcula pesos de área por município
    4. Para cada município, extrai séries de Tmax e Tmin ponderadas
    5. Para cada município-ano, calcula a data de atingimento da soma térmica
    6. Escreve resultados na coluna Antese_v2
    """
    t0 = time.time()

    # ── 1. Planilha ──
    print("[1] Lendo planilha...")
    wb = openpyxl.load_workbook(planilha_path)
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]

    # Mapear colunas
    col_map = {}
    for col_name in [col_municipio, col_ano, col_codigo, col_dec, col_antese_v2]:
        try:
            col_map[col_name] = headers.index(col_name)
        except ValueError:
            if col_name == col_antese_v2:
                print(f"   Coluna '{col_antese_v2}' não encontrada — será criada")
                col_map[col_name] = len(headers)
                ws.cell(row=1, column=col_map[col_name] + 1, value=col_antese_v2)
                headers.append(col_antese_v2)
            else:
                raise ValueError(f"Coluna obrigatória '{col_name}' não encontrada!")

    # Ler dados
    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row[:max(col_map.values()) + 1]):
            continue
        rows_data.append(row)

    print(f"   {len(rows_data)} linhas de dados")

    # Dicionário: (município, ano) → índice em rows_data
    mun_anos = {}
    codigos = {}
    for i, row in enumerate(rows_data):
        mun = str(row[col_map[col_municipio]]).strip() if row[col_map[col_municipio]] else None
        ano = int(row[col_map[col_ano]]) if row[col_map[col_ano]] else None
        cod = row[col_map[col_codigo]]
        if mun and ano:
            mun_anos[(mun, ano)] = i
            codigos[mun] = cod

    print(f"   {len(mun_anos)} municípios-ano únicos")

    # ── 2. Geometrias ──
    print("[2] Carregando geometrias...")
    gdf = gpd.read_file(gpkg_path)
    if "CD_MUN" in gdf.columns:
        gdf["CD_MUN"] = gdf["CD_MUN"].astype(str)
        gdf = gdf.set_index("CD_MUN")

    mun_to_cod = {}
    cod_to_mun = {}
    for mun, cod in codigos.items():
        if cod is not None:
            cod_str = str(int(cod))
            mun_to_cod[mun] = cod_str
            cod_to_mun[cod_str] = mun

    print(f"   {len(mun_to_cod)} municípios mapeados para geometrias")

    # Bbox total com margem
    bounds = gdf.loc[list(mun_to_cod.values())].total_bounds
    MARGEM = 0.2
    LON_MIN = bounds[0] - MARGEM
    LAT_MIN = bounds[1] - MARGEM
    LON_MAX = bounds[2] + MARGEM
    LAT_MAX = bounds[3] + MARGEM
    print(f"   Bbox: lon=[{LON_MIN:.4f}, {LON_MAX:.4f}], "
          f"lat=[{LAT_MIN:.4f}, {LAT_MAX:.4f}]")

    # ── 3. Dados BR-DWGD ──
    print("[3] Carregando dados BR-DWGD (subset)...")
    lats_bbox = np.array([LAT_MIN, LAT_MAX])
    lons_bbox = np.array([LON_MIN, LON_MAX])
    tmax_all, tmin_all, time_idx, lats, lons = carregar_dados_temperatura(
        dir_dados, lats_bbox, lons_bbox
    )

    # Verificar resolução
    lat_res = abs(np.mean(np.diff(lats)))
    lon_res = abs(np.mean(np.diff(lons)))
    assert abs(lat_res - 0.1) < 1e-4, f"Resolução lat {lat_res} ≠ 0.1°"
    assert abs(lon_res - 0.1) < 1e-4, f"Resolução lon {lon_res} ≠ 0.1°"

    # ── 4. Pesos por município ──
    print("[4] Calculando pesos de área...")
    municipios_ok = {}
    sem_geo = []
    sem_inter = []

    for mun_nome, cod_str in mun_to_cod.items():
        if cod_str not in gdf.index:
            sem_geo.append(mun_nome)
            continue

        poly = gdf.loc[cod_str, "geometry"]
        if not poly.is_valid:
            poly = poly.buffer(0)

        weight_map, px_indices = compute_pixel_weights(poly, lats, lons)

        if weight_map is None:
            sem_inter.append(mun_nome)
            continue

        municipios_ok[mun_nome] = {
            "codigo": cod_str,
            "pesos": weight_map,
            "pixels": px_indices,
        }

    print(f"   {len(municipios_ok)} municípios com pesos calculados")
    if sem_geo:
        print(f"   ⚠ Sem geometria: {sem_geo}")
    if sem_inter:
        print(f"   ⚠ Sem interseção: {sem_inter}")

    # ── 5. Extrair séries municipais e calcular GD ──
    print(f"[5] Calculando GD (base={base_temp}°C, alvo={soma_alvo:.0f}°C-dia)...")
    series_gd = {}

    for mun_nome, info in municipios_ok.items():
        weight_map = info["pesos"]

        tmax_mun = np.zeros(len(time_idx), dtype=np.float64)
        tmin_mun = np.zeros(len(time_idx), dtype=np.float64)

        for (i, j), w in weight_map.items():
            tmax_mun += tmax_all[:, i, j] * w
            tmin_mun += tmin_all[:, i, j] * w

        gd_mun = calcular_grau_dia_diario(tmax_mun, tmin_mun, base_temp)
        series_gd[mun_nome] = gd_mun

    print(f"   Séries de GD extraídas para {len(series_gd)} municípios")

    # ── 6. Processar cada município-ano ──
    print(f"[6] Processando {len(mun_anos)} municípios-ano...")

    # Encontrar letra da coluna Antese_v2
    col_letra = None
    for cell in ws[1]:
        if cell.value == col_antese_v2:
            col_letra = cell.column_letter
            break

    cont_atingiu = 0
    cont_nao = 0
    cont_erro = 0
    resultados_log = []

    for (mun_nome, ano), idx_linha in mun_anos.items():
        row = rows_data[idx_linha]
        linha_excel = idx_linha + 2  # 1-based, pulando header

        # Obter data Dec_abril_16mm
        data_dec_str = row[col_map[col_dec]]
        if data_dec_str is None:
            ws[f"{col_letra}{linha_excel}"] = "SEM DATA"
            cont_erro += 1
            continue

        try:
            data_inicio = pd.Timestamp(str(data_dec_str).strip())
        except Exception:
            ws[f"{col_letra}{linha_excel}"] = "DATA_INVALIDA"
            cont_erro += 1
            continue

        if mun_nome not in series_gd:
            ws[f"{col_letra}{linha_excel}"] = "SEM_SERIE"
            cont_erro += 1
            continue

        gd_mun = series_gd[mun_nome]
        data_atinge, soma_atingiu, dias_corridos = encontrar_data_atinge_soma(
            gd_mun, time_idx, data_inicio, soma_alvo
        )

        if data_atinge is not None:
            data_str = data_atinge.strftime("%Y-%m-%d")
            ws[f"{col_letra}{linha_excel}"] = data_str
            cont_atingiu += 1
            resultados_log.append(
                (mun_nome, ano, data_inicio, data_atinge, dias_corridos, soma_atingiu)
            )
        else:
            ws[f"{col_letra}{linha_excel}"] = f"NAO_ATINGIU_{soma_atingiu:.0f}"
            cont_nao += 1

    # ── 7. Salvar ──
    print(f"\n[7] Salvando planilha...")
    wb.save(planilha_path)

    # ── 8. Resumo ──
    print(f"\n{'=' * 60}")
    print(f"RESUMO - {col_antese_v2} (soma térmica {soma_alvo:.0f}°C-dia, "
          f"Tbase={base_temp}°C)")
    print(f"{'=' * 60}")
    print(f"  Total municípios-ano: {len(mun_anos)}")
    print(f"  ✅ Data encontrada:    {cont_atingiu}")
    print(f"  ❌ Não atingiu:        {cont_nao}")
    print(f"  ⚠  Erro:              {cont_erro}")
    print(f"  🌍 Municípios OK:      {len(municipios_ok)}/{len(mun_to_cod)}")
    print(f"\n  Arquivo: {planilha_path}")
    print(f"  Coluna: {col_letra} ({col_antese_v2})")
    print(f"  Tempo total: {time.time() - t0:.1f}s")

    if resultados_log:
        print(f"\n  Amostra (primeiros 10):")
        for m, a, d_ini, d_fim, dias, soma in resultados_log[:10]:
            print(f"    {m} ({a}): ini={d_ini.strftime('%Y-%m-%d')}, "
                  f"fim={d_fim.strftime('%Y-%m-%d')}, "
                  f"dias={dias}, soma={soma:.0f}°C-dia")

    return resultados_log


# ===========================================================================
# CLI
# ===========================================================================

def parse_args():
    p = argparse.ArgumentParser(
        description="Calcula data de atingimento de soma térmica (graus-dia) "
                    "a partir de Tmax/Tmin diários do BR-DWGD.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplo:
  python calcular_antese_v2.py \\
      --planilha PAM_CAFE_MATAS_clima.xlsx \\
      --sheet "DELTA PRODUTIVIDADE (brutos)" \\
      --gpkg MUN_MATAS.gpkg \\
      --dir_dados /tmp/br-dwgd-daily \\
      --soma_alvo 1980.0 \\
      --base_temp 8.5
        """,
    )
    p.add_argument("--planilha", required=True,
                   help="Caminho da planilha Excel (.xlsx)")
    p.add_argument("--sheet", default="DELTA PRODUTIVIDADE (brutos)",
                   help="Nome da aba na planilha")
    p.add_argument("--gpkg", required=True,
                   help="GeoPackage com geometrias dos municípios")
    p.add_argument("--dir_dados", default="/tmp/br-dwgd-daily",
                   help="Diretório com arquivos NetCDF diários do BR-DWGD")
    p.add_argument("--coluna_mun", default="Município",
                   help="Nome da coluna do município")
    p.add_argument("--coluna_ano", default="Ano",
                   help="Nome da coluna do ano")
    p.add_argument("--coluna_cod", default="Código IBGE",
                   help="Nome da coluna do código IBGE")
    p.add_argument("--coluna_dec", default="Dec_abril_16mm",
                   help="Nome da coluna com data de partida")
    p.add_argument("--coluna_antese_v2", default="Antese_v2",
                   help="Nome da coluna alvo para escrita")
    p.add_argument("--soma_alvo", type=float, default=SOMA_ALVO_PADRAO,
                   help=f"Soma térmica alvo em °C-dia (padrão: {SOMA_ALVO_PADRAO})")
    p.add_argument("--base_temp", type=float, default=BASE_TEMP_PADRAO,
                   help=f"Temperatura base em °C (padrão: {BASE_TEMP_PADRAO})")
    return p.parse_args()


def main():
    args = parse_args()
    processar(
        planilha_path=args.planilha,
        sheet_name=args.sheet,
        gpkg_path=args.gpkg,
        dir_dados=args.dir_dados,
        col_municipio=args.coluna_mun,
        col_ano=args.coluna_ano,
        col_codigo=args.coluna_cod,
        col_dec=args.coluna_dec,
        col_antese_v2=args.coluna_antese_v2,
        soma_alvo=args.soma_alvo,
        base_temp=args.base_temp,
    )


if __name__ == "__main__":
    main()
