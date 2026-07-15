#!/usr/bin/env python3
"""
calcular_variavel_periodo.py — Generalizador BR-DWGD
====================================================
Calcula qualquer variável do BR-DWGD (acumulada ou média) para qualquer período
definido por um par de colunas de data em uma planilha Excel, com ponderação
por área de interseção pixel × município.

Resolução espacial obrigatória: 0.1° × 0.1° — nunca reamostrar.

Uso:
    python calcular_variavel_periodo.py \\
        --planilha planilha.xlsx \\
        --sheet "Nome da Sheet" \\
        --gpkg municipios.gpkg \\
        --dir_dados /caminho/DATA_BR_DWGD \\
        --variavel pr \\
        --coluna_mun "Município" \\
        --coluna_cod "Código IBGE" \\
        --coluna_ini "Antese" \\
        --coluna_fim "160dias_antese" \\
        --coluna_alvo "Pr_160d_mm" \\
        --agregacao soma

Variáveis suportadas: pr, Tmax, Tmin, Tmean, Rs, RH, u2, ETo
Agregação sugerida: pr e ETo → soma; demais → media

Funciona com dados DIÁRIOS v3.2.4 (3 arquivos/var) e MENSAIS v3.2.4.

Citação obrigatória:
    Xavier, A. C. et al. (2022). Int. J. Climatol., 42(16), 8390–8404.
    https://doi.org/10.1002/joc.7731
"""

import argparse
import sys
import os
import gc
import time
import warnings
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
import xarray as xr
import geopandas as gpd
import openpyxl
from shapely.geometry import box

warnings.filterwarnings("ignore", category=FutureWarning)

# ── Constantes ──────────────────────────────────────────────────────────────

# Variáveis que usam SOMA como agregação
VARIAVEIS_SOMA = {"pr", "ETo"}
# Variáveis que usam MÉDIA como agregação
VARIAVEIS_MEDIA = {"Tmax", "Tmin", "Tmean", "Rs", "RH", "u2"}
# Todas as variáveis válidas
VARIAVEIS_VALIDAS = VARIAVEIS_SOMA | VARIAVEIS_MEDIA

# Nomes das variáveis nos arquivos NetCDF (v3.2.4)
NOME_NETCDF_VAR = {
    "pr": "pr",
    "ETo": "ETo",
    "Tmax": "Tmax",
    "Tmin": "Tmin",
    "Rs": "Rs",
    "RH": "RH",
    "u2": "u2",
}

# Prefixo dos nomes de arquivo para cada variável
PREFIXO_ARQUIVO = {
    "pr": "pr",
    "ETo": "ETo",
    "Tmax": "Tmax",
    "Tmin": "Tmin",
    "Rs": "Rs",
    "RH": "RH",
    "u2": "u2",
    "Tmean": "Tmean",  # Tmean só existe em mensais
}

# ═══════════════════════════════════════════════════════════════════════════════
# 1. Validação e configuração
# ═══════════════════════════════════════════════════════════════════════════════

def validar_variavel(var: str) -> str:
    """Valida e normaliza o nome da variável."""
    # Aceitar variações de capitalização
    mapa = {
        "pr": "pr", "PR": "pr", "Pr": "pr",
        "eto": "ETo", "ETO": "ETo", "ETo": "ETo",
        "tmax": "Tmax", "TMAX": "Tmax", "Tmax": "Tmax",
        "tmin": "Tmin", "TMIN": "Tmin", "Tmin": "Tmin",
        "tmean": "Tmean", "TMEAN": "Tmean", "Tmean": "Tmean",
        "rs": "Rs", "RS": "Rs", "Rs": "Rs",
        "rh": "RH", "RH": "RH",
        "u2": "u2", "U2": "u2",
    }
    if var in mapa:
        return mapa[var]
    raise ValueError(
        f"Variável '{var}' não reconhecida. Use: {sorted(VARIAVEIS_VALIDAS)}"
    )


def agregacao_padrao(var: str) -> str:
    """Retorna a agregação padrão para a variável."""
    if var in VARIAVEIS_SOMA:
        return "soma"
    return "media"


def validar_agregacao(agregacao: str, variavel: str) -> str:
    """Valida e normaliza o tipo de agregação."""
    agregacao = agregacao.lower().strip()
    if agregacao not in ("soma", "media", "sum", "mean", "avg", "average"):
        raise ValueError("Agregação deve ser 'soma' ou 'media'")
    if agregacao in ("sum",):
        return "soma"
    if agregacao in ("mean", "avg", "average"):
        return "media"
    return agregacao


# ═══════════════════════════════════════════════════════════════════════════════
# 2. Leitura de dados
# ═══════════════════════════════════════════════════════════════════════════════

def ler_planilha(caminho: str, sheet: str) -> tuple:
    """
    Lê a planilha Excel e retorna (workbook, worksheet, headers_map).
    Não carrega tudo em memória — usa openpyxl iterativo.
    """
    wb = openpyxl.load_workbook(caminho)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' não encontrada. Disponíveis: {wb.sheetnames}")
    ws = wb[sheet]
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[str(val).strip()] = col
    return wb, ws, headers


def ler_municipios(gpkg_path: str) -> gpd.GeoDataFrame:
    """Lê o GeoPackage e padroniza a coluna CD_MUN como string."""
    gdf = gpd.read_file(gpkg_path)
    if "CD_MUN" in gdf.columns:
        gdf["CD_MUN"] = gdf["CD_MUN"].astype(str)
    return gdf


def calc_bbox(gdf: gpd.GeoDataFrame, buffer: float = 0.5) -> tuple:
    """Calcula a bounding box expandida dos municípios."""
    bounds = gdf.total_bounds
    return (
        bounds[1] - buffer,  # lat_min
        bounds[3] + buffer,  # lat_max
        bounds[0] - buffer,  # lon_min
        bounds[2] + buffer,  # lon_max
    )


# ═══════════════════════════════════════════════════════════════════════════════
# 3. Pesos de área (pixel × município)
# ═══════════════════════════════════════════════════════════════════════════════

def calcular_pesos_area(
    gdf: gpd.GeoDataFrame,
    lat_region: np.ndarray,
    lon_region: np.ndarray,
) -> dict:
    """
    Calcula os pesos de área para cada município.
    
    Retorna: dict[CD_MUN] -> lista de {lat_idx, lon_idx, weight}
    A soma dos weights para cada município é ~1.0.
    """
    # Criar grid cells
    cells = []
    for i, lat_val in enumerate(lat_region):
        for j, lon_val in enumerate(lon_region):
            cells.append({
                "lat_idx": i,
                "lon_idx": j,
                "geometry": box(lon_val - 0.05, lat_val - 0.05,
                                lon_val + 0.05, lat_val + 0.05),
            })
    grid_gdf = gpd.GeoDataFrame(cells, crs="EPSG:4674")
    
    # Spatial join
    sjoin = gpd.sjoin(gdf, grid_gdf, how="inner", predicate="intersects")
    
    # Calcular pesos
    weights = {}
    for idx_mun in gdf.index:
        cd_mun = gdf.loc[idx_mun, "CD_MUN"]
        mun_geom = gdf.loc[idx_mun, "geometry"]
        total_area = mun_geom.area
        
        rows = sjoin[sjoin.index == idx_mun]
        pixel_weights = []
        for _, row in rows.iterrows():
            cell_geom = grid_gdf.geometry.iloc[row["index_right"]]
            area = mun_geom.intersection(cell_geom).area
            if area > 0:
                pixel_weights.append({
                    "lat_idx": int(row["lat_idx"]),
                    "lon_idx": int(row["lon_idx"]),
                    "weight": area / total_area,
                })
        
        if pixel_weights:
            weights[cd_mun] = pixel_weights
    
    return weights


# ═══════════════════════════════════════════════════════════════════════════════
# 4. Carregamento de dados NetCDF
# ═══════════════════════════════════════════════════════════════════════════════

def encontrar_arquivos(dir_dados: str, prefixo: str) -> list:
    """
    Encontra os arquivos NetCDF para a variável no diretório.
    Suporta v3.2.3 (1 arquivo) e v3.2.4 (3 arquivos).
    Retorna lista ordenada.
    """
    padroes = [f"{prefixo}_", f"{prefixo.lower()}_"]
    arquivos = []
    for fname in sorted(os.listdir(dir_dados)):
        if any(fname.startswith(p) for p in padroes) and fname.endswith(".nc"):
            arquivos.append(os.path.join(dir_dados, fname))
    
    if not arquivos:
        raise FileNotFoundError(
            f"Nenhum arquivo NetCDF encontrado para variável com prefixo '{prefixo}' "
            f"em {dir_dados}"
        )
    return arquivos


def abrir_variavel(
    arquivos: list,
    var_name: str,
    lat_min: float,
    lat_max: float,
    lon_min: float,
    lon_max: float,
    chunks: dict = None,
) -> tuple:
    """
    Abre e concatena arquivos NetCDF de uma variável, com subset espacial.
    
    Args:
        arquivos: lista de caminhos NetCDF (ordenados temporalmente)
        var_name: nome da variável no NetCDF
        lat_min, lat_max: limites de latitude
        lon_min, lon_max: limites de longitude
        chunks: dicionário de chunks (ex: {"time": 365})
    
    Returns:
        (data_array: np.ndarray (time, lat, lon),
         time_index: pd.DatetimeIndex,
         lats: np.ndarray, lons: np.ndarray)
    """
    if chunks is None:
        chunks = {"time": 365}
    
    all_data = []
    all_times = []
    lats = lons = None
    
    for i, nc_path in enumerate(arquivos):
        print(f"    [{i+1}/{len(arquivos)}] {os.path.basename(nc_path)}...", end=" ", flush=True)
        
        with xr.open_dataset(nc_path, chunks=chunks) as ds:
            # Verificar nome da variável
            if var_name not in ds.data_vars:
                # Tentar lowercase
                alt_name = var_name.lower()
                if alt_name in ds.data_vars:
                    var_name = alt_name
                else:
                    # Listar disponíveis
                    raise KeyError(
                        f"Variável '{var_name}' não encontrada em {nc_path}. "
                        f"Disponíveis: {list(ds.data_vars)}"
                    )
            
            # Verificar resolução
            lat_r = abs(float(ds.latitude.diff("latitude").mean()))
            lon_r = abs(float(ds.longitude.diff("longitude").mean()))
            if not (abs(lat_r - 0.1) < 1e-4 and abs(lon_r - 0.1) < 1e-4):
                raise ValueError(
                    f"Resolução {lat_r:.4f}°×{lon_r:.4f}° ≠ 0.1°×0.1° em {nc_path}"
                )
            
            # Verificar ordenação da latitude
            lat_vals = ds.latitude.values
            if lat_vals[0] > lat_vals[-1]:
                # Decrescente: slice(lat_max, lat_min)
                ds_sub = ds.sel(
                    latitude=slice(lat_max, lat_min),
                    longitude=slice(lon_min, lon_max),
                )
            else:
                # Crescente: slice(lat_min, lat_max)
                ds_sub = ds.sel(
                    latitude=slice(lat_min, lat_max),
                    longitude=slice(lon_min, lon_max),
                )
            
            # Carregar dados (após subset espacial)
            data = ds_sub[var_name].values
            times = pd.DatetimeIndex(ds_sub.time.values)
            
            if lats is None:
                lats = ds_sub.latitude.values
                lons = ds_sub.longitude.values
            
            all_data.append(data)
            all_times.append(times)
            
            print(f"OK ({data.shape[0]} dias, {data.shape[1]}×{data.shape[2]} px)")
        
        gc.collect()
    
    # Concatenar
    data_full = np.concatenate(all_data, axis=0).astype(np.float32)
    time_full = all_times[0]
    for t in all_times[1:]:
        time_full = time_full.append(t)
    
    return data_full, time_full, lats, lons


# ═══════════════════════════════════════════════════════════════════════════════
# 5. Extração municipal ponderada
# ═══════════════════════════════════════════════════════════════════════════════

def extrair_series_municipais(
    data: np.ndarray,
    time_idx: pd.DatetimeIndex,
    weights: dict,
    progress_desc: str = "",
) -> dict:
    """
    Extrai séries temporais ponderadas para cada município.
    
    Args:
        data: array (time, lat, lon)
        time_idx: índice temporal
        weights: dict[CD_MUN] -> lista de pesos
        progress_desc: descrição para print
    
    Returns:
        dict[CD_MUN] -> array 1D (time)
    """
    series = {}
    n_mun = len(weights)
    
    for n, (cd_mun, cells) in enumerate(weights.items()):
        if progress_desc and (n % 10 == 0 or n == n_mun - 1):
            print(f"    {progress_desc}: {n+1}/{n_mun}", end="\r", flush=True)
        
        vals = np.zeros(len(time_idx), dtype=np.float64)
        for cell in cells:
            vals += data[:, cell["lat_idx"], cell["lon_idx"]] * cell["weight"]
        series[cd_mun] = vals.astype(np.float32)
    
    if progress_desc:
        print(f"    {progress_desc}: {n_mun}/{n_mun} concluído")
    
    return series


# ═══════════════════════════════════════════════════════════════════════════════
# 6. Cálculo do período
# ═══════════════════════════════════════════════════════════════════════════════

def calcular_por_periodo(
    series: dict,
    time_idx: pd.DatetimeIndex,
    df_datas: list,
    agregacao: str,
    peso_diario: float = 1.0,
) -> np.ndarray:
    """
    Calcula soma ou média para cada registro entre duas datas.
    
    Args:
        series: dict[CD_MUN] -> array 1D (time)
        time_idx: índice temporal global
        df_datas: lista de (cd_mun, data_ini, data_fim)
        agregacao: "soma" ou "media"
        peso_diario: fator de peso para cada dia (ex: ETo diário em mm/dia → 1.0)
    
    Returns:
        array com resultados (NaN se não encontrado)
    """
    # Mapa: data YYYY-MM-DD -> índice
    date_to_idx = {d.strftime("%Y-%m-%d"): i for i, d in enumerate(time_idx)}
    
    resultados = np.full(len(df_datas), np.nan, dtype=np.float64)
    n_ok = 0
    n_fora = 0
    n_sem_mun = 0
    
    for i, (cd_mun, d0, d1) in enumerate(df_datas):
        if cd_mun not in series:
            n_sem_mun += 1
            continue
        
        d0_str = d0.strftime("%Y-%m-%d")
        d1_str = d1.strftime("%Y-%m-%d")
        
        if d0_str not in date_to_idx or d1_str not in date_to_idx:
            n_fora += 1
            continue
        
        idx0 = date_to_idx[d0_str]
        idx1 = date_to_idx[d1_str]
        
        segmento = series[cd_mun][idx0:idx1 + 1] * peso_diario
        
        if agregacao == "soma":
            resultados[i] = float(np.sum(segmento))
        else:  # media
            resultados[i] = float(np.mean(segmento))
        
        n_ok += 1
    
    if n_fora > 0:
        print(f"    ⚠ {n_fora} registros com data fora do período dos dados")
    if n_sem_mun > 0:
        print(f"    ⚠ {n_sem_mun} registros sem série municipal")
    
    return resultados, n_ok


# ═══════════════════════════════════════════════════════════════════════════════
# 7. Escrita na planilha
# ═══════════════════════════════════════════════════════════════════════════════

def escrever_resultados(
    ws,
    col_alvo: int,
    resultados: np.ndarray,
    linha_inicio: int = 2,
) -> int:
    """Escreve os resultados na coluna alvo da planilha."""
    n_escritos = 0
    for i, val in enumerate(resultados):
        if not np.isnan(val):
            ws.cell(row=linha_inicio + i, column=col_alvo, value=round(float(val), 2))
            n_escritos += 1
    return n_escritos


# ═══════════════════════════════════════════════════════════════════════════════
# 8. Pipeline principal
# ═══════════════════════════════════════════════════════════════════════════════

def pipeline(
    planilha: str,
    sheet: str,
    gpkg: str,
    dir_dados: str,
    variavel: str,
    coluna_mun: str,
    coluna_cod: str,
    coluna_ini: str,
    coluna_fim: str,
    coluna_alvo: str,
    agregacao: str = None,
    buffer: float = 0.5,
    formato: str = "diario",
):
    """
    Pipeline completo:
    1. Lê planilha e geometrias
    2. Encontra e carrega dados BR-DWGD (subset espacial)
    3. Calcula pesos de área
    4. Extrai séries municipais
    5. Calcula soma/média no período
    6. Escreve na planilha
    """
    t0 = time.time()
    
    # ── 0. Validações ──
    variavel = validar_variavel(variavel)
    if agregacao is None:
        agregacao = agregacao_padrao(variavel)
    else:
        agregacao = validar_agregacao(agregacao, variavel)
    
    var_netcdf = NOME_NETCDF_VAR.get(variavel, variavel)
    prefixo = PREFIXO_ARQUIVO.get(variavel, variavel)
    
    print(f"\n{'='*65}")
    print(f"📊 BR-DWGD — Cálculo Generalizado")
    print(f"{'='*65}")
    print(f"  Variável:     {variavel} → '{var_netcdf}' no NetCDF")
    print(f"  Agregação:    {agregacao.upper()}")
    print(f"  Período:      '{coluna_ini}' → '{coluna_fim}'")
    print(f"  Coluna alvo:  '{coluna_alvo}'")
    print(f"  Planilha:     {os.path.basename(planilha)} @ [{sheet}]")
    print(f"  Formato:      {formato}")
    print(f"{'='*65}\n")
    
    # ── 1. Planilha ──
    print("[1/7] Lendo planilha...")
    wb, ws, headers = ler_planilha(planilha, sheet)
    
    for nome_col in [coluna_mun, coluna_cod, coluna_ini, coluna_fim, coluna_alvo]:
        if nome_col not in headers:
            raise KeyError(f"Coluna '{nome_col}' não encontrada na planilha")
    
    col_ini = headers[coluna_ini]
    col_fim = headers[coluna_fim]
    col_cod = headers[coluna_cod]
    col_alvo = headers[coluna_alvo]
    
    print(f"  Colunas: ini={col_ini}, fim={col_fim}, cod={col_cod}, alvo={col_alvo}")
    print(f"  Linhas: {ws.max_row - 1} registros")
    
    # Coletar datas e códigos
    registros = []
    n_linhas = ws.max_row - 1
    for r in range(2, ws.max_row + 1):
        cod = str(ws.cell(row=r, column=col_cod).value or "").strip()
        v_ini = ws.cell(row=r, column=col_ini).value
        v_fim = ws.cell(row=r, column=col_fim).value
        
        # Pular linhas sem dados
        if not cod or cod == "None" or v_ini is None or v_fim is None:
            registros.append((cod, None, None))
            continue
        
        # Converter datas
        def parse_date(v):
            if isinstance(v, datetime):
                return v
            if isinstance(v, str):
                try:
                    return datetime.strptime(v.strip()[:10], "%Y-%m-%d")
                except ValueError:
                    try:
                        return datetime.strptime(v.strip()[:10], "%d/%m/%Y")
                    except ValueError:
                        return None
            return None
        
        d_ini = parse_date(v_ini)
        d_fim = parse_date(v_fim)
        
        if d_ini is None or d_fim is None:
            registros.append((cod, None, None))
            continue
        
        registros.append((cod, d_ini, d_fim))
    
    n_validos = sum(1 for _, d0, d1 in registros if d0 is not None and d1 is not None)
    print(f"  Registros com datas válidas: {n_validos}/{n_linhas}")
    
    if n_validos == 0:
        print("  ❌ Nenhum registro com data válida. Abortando.")
        wb.close()
        return
    
    # ── 2. Geometrias ──
    print("[2/7] Carregando geometrias municipais...")
    gdf = ler_municipios(gpkg)
    lat_min, lat_max, lon_min, lon_max = calc_bbox(gdf, buffer)
    print(f"  {len(gdf)} municípios, bbox: [{lon_min:.2f}, {lon_max:.2f}] × [{lat_min:.2f}, {lat_max:.2f}]")
    
    # ── 3. Encontrar arquivos ──
    print("[3/7] Localizando arquivos NetCDF...")
    arquivos = encontrar_arquivos(dir_dados, prefixo)
    print(f"  Encontrados: {len(arquivos)} arquivo(s)")
    for f in arquivos:
        sz = os.path.getsize(f) / 1e9
        print(f"    {os.path.basename(f)} ({sz:.2f} GB)")
    
    # ── 4. Carregar dados ──
    print("[4/7] Carregando dados...")
    if formato == "diario":
        chunks = {"time": 365}
    else:
        chunks = {"time": 120}
    
    data, time_idx, lats_region, lons_region = abrir_variavel(
        arquivos, var_netcdf, lat_min, lat_max, lon_min, lon_max, chunks
    )
    print(f"  Dados carregados: {data.shape[0]} dias × {data.shape[1]} lat × {data.shape[2]} lon")
    print(f"  Período: {time_idx[0].strftime('%Y-%m-%d')} a {time_idx[-1].strftime('%Y-%m-%d')}")
    print(f"  Memória: {data.nbytes / 1e6:.0f} MB")
    
    # ── 5. Pesos de área ──
    print("[5/7] Calculando pesos de área...")
    weights = calcular_pesos_area(gdf, lats_region, lons_region)
    print(f"  Pesos calculados para {len(weights)} municípios")
    
    # Verificar quais municípios da planilha estão nos pesos
    cods_planilha = set(r[0] for r in registros if r[0])
    cods_pesos = set(weights.keys())
    cods_faltando = cods_planilha - cods_pesos
    if cods_faltando:
        print(f"  ⚠ {len(cods_faltando)} municípios na planilha sem geometria: {sorted(cods_faltando)[:5]}...")
    
    # ── 6. Extrair séries municipais ──
    print("[6/7] Extraindo séries municipais ponderadas...")
    series = extrair_series_municipais(data, time_idx, weights, progress_desc="  Municípios")
    
    # ── 7. Calcular período ──
    print(f"[7/7] Calculando {agregacao} no período '{coluna_ini}'→'{coluna_fim}'...")
    
    # Determinar peso diário (para ETo diário em mm/dia, o valor já é diário)
    peso_diario = 1.0  # Padrão: dados já estão na unidade correta por dia
    
    df_datas = [(r[0], r[1], r[2]) for r in registros if r[1] is not None and r[2] is not None]
    
    resultados, n_ok = calcular_por_periodo(
        series, time_idx, df_datas, agregacao, peso_diario
    )
    
    # ── Escrever na planilha ──
    print(f"\n  Escrevendo {n_ok} resultados na coluna '{coluna_alvo}'...")
    # Mapear resultados de volta para linhas
    idx_resultado = 0
    n_escritos = 0
    for r in range(2, ws.max_row + 1):
        cod = str(ws.cell(row=r, column=col_cod).value or "").strip()
        v_ini = ws.cell(row=r, column=col_ini).value
        v_fim = ws.cell(row=r, column=col_fim).value
        if cod and cod != "None" and v_ini is not None and v_fim is not None:
            if idx_resultado < len(resultados) and not np.isnan(resultados[idx_resultado]):
                ws.cell(row=r, column=col_alvo, value=round(float(resultados[idx_resultado]), 2))
                n_escritos += 1
            idx_resultado += 1
    
    print(f"  ✅ {n_escritos} valores escritos com sucesso!")
    
    # ── Salvar ──
    print(f"\n  Salvando planilha...")
    wb.save(planilha)
    wb.close()
    print(f"  ✅ Planilha salva: {os.path.basename(planilha)}")
    
    # ── Estatísticas ──
    vals_validos = resultados[~np.isnan(resultados)]
    if len(vals_validos) > 0:
        print(f"\n{'─'*50}")
        print(f"📈 Estatísticas — {coluna_alvo} ({variavel})")
        print(f"{'─'*50}")
        print(f"  Registros:  {len(vals_validos)}")
        print(f"  Média:      {np.mean(vals_validos):.2f}")
        print(f"  Mediana:    {np.median(vals_validos):.2f}")
        print(f"  Mín:        {np.min(vals_validos):.2f}")
        print(f"  Máx:        {np.max(vals_validos):.2f}")
        print(f"  Desv. Pad.: {np.std(vals_validos):.2f}")
        print(f"{'─'*50}")
    
    elapsed = time.time() - t0
    print(f"\n⏱  Tempo total: {elapsed/60:.1f} min")
    print(f"{'='*65}")
    print(f"✅ Processamento concluído!")
    print(f"{'='*65}")


# ═══════════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════════

def parse_args():
    p = argparse.ArgumentParser(
        description="Calcula variáveis do BR-DWGD para períodos definidos em planilha",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos:
  # Precipitação acumulada entre Antese e 160dias_antese
  python calcular_variavel_periodo.py --planilha dados.xlsx --sheet "Sheet1" \\
      --gpkg municipios.gpkg --dir_dados ../DATA_BR_DWGD --variavel pr \\
      --coluna_mun "Município" --coluna_cod "Código IBGE" \\
      --coluna_ini "Antese" --coluna_fim "160dias_antese" --coluna_alvo "Pr_160d_mm"
  
  # ETo média entre duas datas
  python calcular_variavel_periodo.py --planilha dados.xlsx --sheet "Sheet1" \\
      --gpkg municipios.gpkg --dir_dados ../DATA_BR_DWGD --variavel ETo \\
      --coluna_ini "Data1" --coluna_fim "Data2" --coluna_alvo "ETo_media" \\
      --agregacao media
  
  # Temperatura máxima média no período
  python calcular_variavel_periodo.py --planilha dados.xlsx --sheet "Sheet1" \\
      --gpkg municipios.gpkg --dir_dados ../DATA_BR_DWGD --variavel Tmax \\
      --coluna_ini "Inicio" --coluna_fim "Fim" --coluna_alvo "Tmax_media" \\
      --agregacao media
        """,
    )
    
    p.add_argument("--planilha", required=True, help="Caminho para planilha Excel (.xlsx)")
    p.add_argument("--sheet", default=None, help="Nome da sheet (opcional se só houver uma)")
    p.add_argument("--gpkg", required=True, help="Caminho para GeoPackage dos municípios")
    p.add_argument("--dir_dados", required=True, help="Diretório com arquivos NetCDF do BR-DWGD")
    p.add_argument("--variavel", required=True, help=f"Variável: {sorted(VARIAVEIS_VALIDAS)}")
    
    p.add_argument("--coluna_mun", default="Município", help="Nome da coluna com nome do município")
    p.add_argument("--coluna_cod", default="Código IBGE", help="Nome da coluna com código IBGE")
    p.add_argument("--coluna_ini", required=True, help="Nome da coluna com data inicial")
    p.add_argument("--coluna_fim", required=True, help="Nome da coluna com data final")
    p.add_argument("--coluna_alvo", required=True, help="Nome da coluna para escrever resultado")
    
    p.add_argument("--agregacao", default=None,
                   help="'soma' (acumulado) ou 'media' (média). Padrão: auto (pr/ETo=soma, demais=media)")
    p.add_argument("--buffer", type=float, default=0.5,
                   help="Buffer em graus ao redor dos municípios (padrão: 0.5)")
    p.add_argument("--formato", choices=["diario", "mensal"], default="diario",
                   help="Formato dos dados (padrão: diario)")
    
    return p.parse_args()


def main():
    args = parse_args()
    
    # Detectar sheet se não especificada
    sheet = args.sheet
    if sheet is None:
        import openpyxl
        wb = openpyxl.load_workbook(args.planilha, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        if len(sheets) == 1:
            sheet = sheets[0]
            print(f"Sheet não especificada. Usando única sheet: '{sheet}'")
        else:
            print(f"Múltiplas sheets: {sheets}")
            print("Use --sheet para especificar.")
            sys.exit(1)
    
    pipeline(
        planilha=args.planilha,
        sheet=sheet,
        gpkg=args.gpkg,
        dir_dados=args.dir_dados,
        variavel=args.variavel,
        coluna_mun=args.coluna_mun,
        coluna_cod=args.coluna_cod,
        coluna_ini=args.coluna_ini,
        coluna_fim=args.coluna_fim,
        coluna_alvo=args.coluna_alvo,
        agregacao=args.agregacao,
        buffer=args.buffer,
        formato=args.formato,
    )


if __name__ == "__main__":
    main()
